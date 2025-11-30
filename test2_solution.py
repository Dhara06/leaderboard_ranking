import pandas as pd
import numpy as np
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font

xls = pd.ExcelFile("leaderboard.xlsx")

df = pd.read_excel(xls)
scores_df = df.iloc[:26,:]
spend_df = df.iloc[31:59,:]

scores_df.columns = scores_df.iloc[0]
scores_df = scores_df[1:].reset_index(drop=True)

spend_df.columns = spend_df.iloc[0]
spend_df = spend_df[1:].reset_index(drop=True)

new_cols = []
pts_count = 1
for col in scores_df.columns:
    if col == "Pts":
        new_cols.append(f"R{pts_count}")
        pts_count += 1
    else:
        new_cols.append(col)

scores_df.columns = new_cols

point_columns = [c for c in scores_df.columns if c.startswith("R")]
scores_df[point_columns] = (scores_df[point_columns].replace(["-", "D$Q"], 0).apply(pd.to_numeric, errors="coerce").fillna(0))

scores_df["Total_Points"] = scores_df[point_columns].sum(axis=1)

spend_event_cols = [col for col in spend_df.columns if isinstance(col, str) and "$" in col]

spend_df[spend_event_cols] = (spend_df[spend_event_cols].replace(["-", "D$Q"], 0).apply(pd.to_numeric, errors="coerce").fillna(0))

scores_df["Total_Spent"] = spend_df[spend_event_cols].sum(axis=1)

def countback_vector(row):
    scores = sorted(row[point_columns], reverse=True)
    counter = Counter(scores)

    vector = []
    for score in sorted(counter.keys(), reverse=True):
        vector.append(score)
        vector.append(counter[score])

    return tuple(vector)

scores_df["Countback"] = scores_df.apply(countback_vector, axis=1)

scores_df_sorted = scores_df.sort_values(
    by=["Total_Points", "Total_Spent", "Countback", "Player"],
    ascending=[False, True, False, True]
)

scores_df_sorted["Pos"] = np.arange(1, len(scores_df_sorted) + 1)

tie_cols = ["Total_Points", "Total_Spent", "Countback"]
scores_df_sorted["Is_Tied"] = scores_df_sorted.duplicated(subset=tie_cols, keep=False)

output_file = "final_sorted_leaderboard.xlsx"
scores_df_sorted.to_excel(output_file, index=False)


wb = load_workbook(output_file)
ws = wb.active

for row in range(2, ws.max_row + 1):
    is_tied_col = ws.max_column
    if ws.cell(row=row, column=is_tied_col).value:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = Font(color="FF0000")

wb.save(output_file)

print("Final sorted leaderboard generated:", output_file)

